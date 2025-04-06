from flask import Flask, render_template, request, redirect, url_for, session, jsonify, send_file
from flask_socketio import SocketIO, emit
from functools import wraps
import json
import os
import logging
import subprocess
import signal
import shutil
import time
import threading
from datetime import datetime, timedelta
import platform
import sys
import psutil
from dotenv import load_dotenv
import uuid
import tempfile
import re
import random
import werkzeug.utils
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import PatternFill

# Monkey patch for gevent compatibility with Python 3.12
# We're not using WebSockets, so we don't need gevent's WebSocket support
from gevent import monkey
monkey.patch_all(ssl=False)

# Load environment variables
load_dotenv()

app = Flask(__name__)
# Get secret key from environment variable
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'default_secret_key')
# Initialize SocketIO with gevent
socketio = SocketIO(app, cors_allowed_origins="*",
                    async_mode='gevent', logger=True, engineio_logger=True)
logging.basicConfig(level=logging.DEBUG)

# Create uploads directory if it doesn't exist
UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Configure file uploads
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create pics directory if it doesn't exist
PICS_FOLDER = 'pics'
if not os.path.exists(PICS_FOLDER):
    os.makedirs(PICS_FOLDER)

# Configure pics upload
ALLOWED_IMAGE_EXTENSIONS = {'jpg', 'jpeg', 'png'}

# Add these global variables at the top level
bot_process = None
bot_connected = False
# Flag to control background threads
should_run_background_tasks = True

# Get dashboard credentials from environment variables
DASHBOARD_USERNAME = os.environ.get('DASHBOARD_USERNAME', 'bot')
DASHBOARD_PASSWORD = os.environ.get('DASHBOARD_PASSWORD', 'bot-bot')

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Helper function to check allowed file extensions


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Helper function to validate phone number


def is_valid_phone_number(number):
    # Remove plus sign if present
    if number.startswith('+'):
        number = number[1:]
    # Check if the number contains only digits and has a reasonable length
    return number.isdigit() and 8 <= len(number) <= 15


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if username == DASHBOARD_USERNAME and password == DASHBOARD_PASSWORD:
            session['logged_in'] = True
            session['username'] = username
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Invalid credentials')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))


@app.route('/')
@login_required
def index():
    # Get username from session or use default
    username = session.get('username', 'admin')
    user = {'username': username}
    return render_template('index.html', user=user)


@app.route('/get_qr_code')
@login_required
def get_qr_code():
    qr_code_file = 'qr_code.png'
    if os.path.exists(qr_code_file):
        return send_file(qr_code_file, mimetype='image/png')
    return jsonify({"message": "QR code not available"})


@app.route('/bot_status')
@login_required
def bot_status():
    global bot_connected
    return jsonify({"connected": bot_connected})


@app.route('/is_bot_ready')
@login_required
def is_bot_ready():
    global bot_connected
    return jsonify({"ready": bot_connected})


@app.route('/qr_code_exists')
@login_required
def qr_code_exists():
    global bot_process, bot_connected

    # Check if bot is running but not connected (connecting state)
    bot_running = bot_process is not None and bot_process.poll() is None
    is_connecting = bot_running and not bot_connected

    # Only check for QR code if we're in a connecting state
    qr_exists = os.path.exists('qr_code.png') if is_connecting else False

    return jsonify({
        "exists": qr_exists,
        "is_connecting": is_connecting,
        "bot_running": bot_running
    })


@app.route('/set_bot_connected', methods=['POST'])
def set_bot_connected():
    global bot_connected
    bot_connected = True
    if os.path.exists('qr_code.png'):
        os.remove('qr_code.png')
    return jsonify({"message": "Bot connection status updated", "ready": True})


@app.route('/reset_bot')
@login_required
def reset_bot():
    global bot_process, bot_connected

    # Update UI immediately
    socketio.emit('bot_status', {'connected': False, 'status': 'resetting'})

    # Stop the bot if it's running
    if bot_process is not None and bot_process.poll() is None:
        try:
            os.kill(bot_process.pid, signal.SIGTERM)
            bot_process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            # Force kill if it doesn't terminate gracefully
            os.kill(bot_process.pid, signal.SIGKILL)
        except Exception as e:
            app.logger.error(f"Error stopping bot during reset: {str(e)}")

        bot_process = None

    # Clear the auth directory
    cache_dir = '.wwebjs_auth'
    if os.path.exists(cache_dir):
        max_attempts = 5
        for attempt in range(max_attempts):
            try:
                shutil.rmtree(cache_dir)
                break
            except PermissionError:
                if attempt < max_attempts - 1:
                    time.sleep(1)
                else:
                    return jsonify({
                        "message": "Failed to remove cache. Please try again.",
                        "connected": False,
                        "error": True,
                        "status": "error"
                    })

    # Remove old QR code if it exists
    if os.path.exists('qr_code.png'):
        try:
            os.remove('qr_code.png')
        except PermissionError:
            pass

    # Start the bot
    try:
        bot_process = subprocess.Popen(['node', 'index.js'])

        # Wait a moment to check if the process started successfully
        time.sleep(1)
        if bot_process.poll() is not None:
            # Process terminated immediately
            return jsonify({
                "message": "Failed to restart bot. Check server logs for details.",
                "connected": False,
                "error": True,
                "status": "error"
            })

        bot_connected = False

        # Emit status update via WebSocket
        socketio.emit(
            'bot_status', {'connected': False, 'status': 'connecting'})

        return jsonify({
            "message": "Bot reset successfully. Please wait for the QR code to appear.",
            "connected": False,
            "status": "connecting"
        })
    except Exception as e:
        app.logger.error(f"Error restarting bot after reset: {str(e)}")
        return jsonify({
            "message": f"Error restarting bot: {str(e)}",
            "connected": False,
            "error": True,
            "status": "error"
        })


@app.route('/set_bot_disconnected', methods=['POST'])
def set_bot_disconnected():
    global bot_connected
    bot_connected = False
    return jsonify({'message': 'Bot disconnected status updated'})


@app.route('/start_bot')
@login_required
def start_bot():
    global bot_process, bot_connected

    # If bot is already running, return success
    if bot_process is not None and bot_process.poll() is None:
        return jsonify({
            "message": "Bot is already running",
            "connected": bot_connected,
            "status": "connecting" if not bot_connected else "connected"
        })

    # Check if auth directory exists and has session data
    auth_dir = '.wwebjs_auth'
    session_exists = os.path.exists(auth_dir) and os.path.exists(
        os.path.join(auth_dir, 'session'))

    # Start the bot
    try:
        bot_process = subprocess.Popen(['node', 'index.js'])

        # Wait a moment to check if the process started successfully
        time.sleep(1)
        if bot_process.poll() is not None:
            # Process terminated immediately
            return jsonify({
                "message": "Failed to start bot. Check server logs for details.",
                "connected": False,
                "error": True,
                "status": "error"
            })

        # Emit status update via WebSocket
        socketio.emit(
            'bot_status', {'connected': False, 'status': 'connecting'})

        # If we have a session, we might reconnect automatically
        connection_message = "Bot started successfully. Attempting to reconnect to existing session..." if session_exists else "Bot started successfully. Waiting for connection..."

        return jsonify({
            "message": connection_message,
            "connected": False,
            "status": "connecting",
            "session_exists": session_exists
        })
    except Exception as e:
        app.logger.error(f"Error starting bot: {str(e)}")
        return jsonify({
            "message": f"Error starting bot: {str(e)}",
            "connected": False,
            "error": True,
            "status": "error"
        })


@app.route('/stop_bot')
@login_required
def stop_bot():
    global bot_process, bot_connected

    # If bot is running, stop it
    if bot_process is not None and bot_process.poll() is None:
        try:
            os.kill(bot_process.pid, signal.SIGTERM)
            # Wait up to 5 seconds for process to terminate
            bot_process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            # Force kill if it doesn't terminate gracefully
            os.kill(bot_process.pid, signal.SIGKILL)
        except Exception as e:
            app.logger.error(f"Error stopping bot: {str(e)}")
            return jsonify({
                "message": f"Error stopping bot: {str(e)}",
                "connected": bot_connected,
                "error": True
            })

        bot_process = None

    # Update connection status
    bot_connected = False

    # Emit status update via WebSocket
    socketio.emit('bot_status', {'connected': False, 'status': 'stopped'})

    return jsonify({
        "message": "Bot stopped successfully",
        "connected": False,
        "status": "stopped"
    })


@app.route('/system_info')
@login_required
def system_info():
    """Endpoint to provide system information for the dashboard."""
    return jsonify(get_system_info())


@app.route('/login.js')
def serve_login_js():
    """Serve the login.js file."""
    return send_file('login.js', mimetype='application/javascript')


# Excel Bulk Messaging Routes
@app.route('/upload_excel', methods=['POST'])
@login_required
def upload_excel():
    if not bot_connected:
        return jsonify({
            "success": False,
            "message": "WhatsApp bot is not connected. Please connect the bot first."
        })

    # Check if file was included in the request
    if 'excel_file' not in request.files:
        return jsonify({
            "success": False,
            "message": "No file provided"
        })

    file = request.files['excel_file']

    # Check if a file was selected
    if file.filename == '':
        return jsonify({
            "success": False,
            "message": "No file selected"
        })

    # Check if the file is allowed
    if not allowed_file(file.filename):
        return jsonify({
            "success": False,
            "message": "Invalid file format. Only Excel files (.xlsx, .xls) are allowed."
        })

    # Generate a unique filename to prevent overwriting
    filename = str(uuid.uuid4()) + '_' + secure_filename(file.filename)
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        # Save the uploaded file
        file.save(file_path)

        # Process the Excel file to validate its structure
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Check if the file has the required structure
            if sheet.max_column < 2:
                os.remove(file_path)  # Clean up
                return jsonify({
                    "success": False,
                    "message": "Invalid Excel format. The file must have at least 2 columns (Name and Phone Number)."
                })

            # Count valid numbers
            total_numbers = 0
            processed_numbers = 0
            invalid_numbers = 0

            # Start from the second row (skip header)
            for row in range(2, sheet.max_row + 1):
                phone_number = str(sheet.cell(row=row, column=2).value or "")
                status = str(sheet.cell(
                    row=row, column=3).value or "").strip().lower()

                if phone_number:
                    total_numbers += 1

                    # Check if this number has already been processed
                    if status in ["success", "fail", "number doesn't exist on whatsapp"]:
                        processed_numbers += 1

                    # Check if phone number is valid
                    if not is_valid_phone_number(phone_number):
                        invalid_numbers += 1

            # Save the processed file
            workbook.save(file_path)

            # Return success response with session ID and file info
            return jsonify({
                "success": True,
                "message": "File uploaded successfully",
                "filename": filename,
                "total_numbers": total_numbers,
                "processed_numbers": processed_numbers,
                "remaining_numbers": total_numbers - processed_numbers,
                "invalid_numbers": invalid_numbers
            })

        except Exception as e:
            # If there's an error processing the Excel file, remove it and return error
            if os.path.exists(file_path):
                os.remove(file_path)
            return jsonify({
                "success": False,
                "message": f"Error processing Excel file: {str(e)}"
            })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error saving file: {str(e)}"
        })


@app.route('/start_bulk_messaging', methods=['POST'])
@login_required
def start_bulk_messaging():
    if not bot_connected:
        return jsonify({
            "success": False,
            "message": "WhatsApp bot is not connected. Please connect the bot first."
        })

    data = request.json
    filename = data.get('filename')
    message_text = data.get('message', '')
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    # Validate inputs
    if not filename or not os.path.exists(file_path):
        return jsonify({
            "success": False,
            "message": "Invalid file or file not found"
        })

    if not message_text:
        return jsonify({
            "success": False,
            "message": "Message text is required"
        })

    # Check if image was uploaded
    has_image = data.get('has_image', False)
    image_data = data.get('image_data', None)

    # Start a background thread to process the messages
    thread = threading.Thread(
        target=process_bulk_messages,
        args=(file_path, message_text, has_image, image_data)
    )
    thread.daemon = True
    thread.start()

    return jsonify({
        "success": True,
        "message": "Bulk messaging started in the background"
    })


@app.route('/download_excel/<filename>')
@login_required
def download_excel(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    # Check if file exists
    if not os.path.exists(file_path):
        return jsonify({
            "success": False,
            "message": "File not found"
        }), 404

    # Return the file for download
    return send_file(file_path, as_attachment=True)


@app.route('/get_progress/<filename>')
@login_required
def get_progress(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    # Check if file exists
    if not os.path.exists(file_path):
        return jsonify({
            "success": False,
            "message": "File not found"
        }), 404

    try:
        # Load the Excel file to check progress
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        total_numbers = 0
        processed_numbers = 0
        success_count = 0
        fail_count = 0
        not_on_whatsapp_count = 0

        # Start from the second row (skip header)
        for row in range(2, sheet.max_row + 1):
            phone_number = str(sheet.cell(row=row, column=2).value or "")
            status = str(sheet.cell(
                row=row, column=3).value or "").strip().lower()

            if phone_number:
                total_numbers += 1

                if status == "success":
                    processed_numbers += 1
                    success_count += 1
                elif status == "fail":
                    processed_numbers += 1
                    fail_count += 1
                elif status in ["number doesn't exist on whatsapp", "number doesn't exist on whatsapp"]:
                    processed_numbers += 1
                    not_on_whatsapp_count += 1

        return jsonify({
            "success": True,
            "total_numbers": total_numbers,
            "processed_numbers": processed_numbers,
            "success_count": success_count,
            "fail_count": fail_count,
            "not_on_whatsapp_count": not_on_whatsapp_count,
            "progress_percentage": (processed_numbers / total_numbers * 100) if total_numbers > 0 else 0
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Error getting progress: {str(e)}"
        })


def process_bulk_messages(file_path, message_text, has_image, image_data):
    try:
        # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Define status cell colors
        success_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE",
                                end_color="FFC7CE", fill_type="solid")
        not_on_whatsapp_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

        # Create a temporary file for the image if provided
        temp_image_path = None
        if has_image and image_data:
            # Remove the data:image/jpeg;base64, prefix
            image_data = image_data.split(
                ',')[1] if ',' in image_data else image_data

            # Create a temporary file
            temp_fd, temp_image_path = tempfile.mkstemp(suffix='.jpg')
            os.close(temp_fd)

            # Write the decoded image data to the temporary file
            with open(temp_image_path, 'wb') as f:
                import base64
                f.write(base64.b64decode(image_data))

        # Process each row in the Excel file
        for row in range(2, sheet.max_row + 1):
            # Get the name and phone number
            name = str(sheet.cell(row=row, column=1).value or "")
            phone_number = str(sheet.cell(row=row, column=2).value or "")
            status = str(sheet.cell(
                row=row, column=3).value or "").strip().lower()

            # Skip rows that have already been processed
            if status in ["success", "fail", "number doesn't exist on whatsapp"]:
                continue

            # Skip empty phone numbers
            if not phone_number:
                continue

            # Format the phone number (remove + if present)
            if phone_number.startswith('+'):
                phone_number = phone_number[1:]

            # Replace placeholders in the message
            personalized_message = message_text.replace('{name}', name)

            # Send the message
            try:
                # Create payload for sending message
                if has_image and temp_image_path:
                    # Run JavaScript to send message with image
                    send_script = f"""
                    const {MessageMedia} = require('whatsapp-web.js');
                    const fs = require('fs');
                    
                    (async () => {{
                        try {{
                            // Check if number exists on WhatsApp
                            const isRegistered = await client.isRegisteredUser('{phone_number}@c.us');
                            if (!isRegistered) {{
                                console.log('NUMBER_NOT_ON_WHATSAPP');
                                return;
                            }}
                            
                            // Send the image with caption
                            const media = MessageMedia.fromFilePath('{temp_image_path.replace('\\', '\\\\')}');
                            await client.sendMessage('{phone_number}@c.us', media, {{ caption: `{personalized_message}` }});
                            console.log('SUCCESS');
                        }} catch(error) {{
                            console.log('ERROR: ' + error.message);
                        }}
                    }})();
                    """

                    # Write the script to a temporary file
                    temp_script_fd, temp_script_path = tempfile.mkstemp(
                        suffix='.js')
                    os.close(temp_script_fd)
                    with open(temp_script_path, 'w') as f:
                        f.write(send_script)

                    # Run the script with Node.js
                    process = subprocess.Popen(
                        ['node', temp_script_path],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True
                    )
                    stdout, stderr = process.communicate()

                    # Clean up the temporary script
                    os.remove(temp_script_path)

                    # Check the result
                    if 'SUCCESS' in stdout:
                        status = "Success"
                        sheet.cell(row=row, column=3).fill = success_fill
                    elif 'NUMBER_NOT_ON_WHATSAPP' in stdout:
                        status = "Number Doesn't Exist on WhatsApp"
                        sheet.cell(
                            row=row, column=3).fill = not_on_whatsapp_fill
                    else:
                        status = "Fail"
                        sheet.cell(row=row, column=3).fill = fail_fill
                else:
                    # Run JavaScript to send text message only
                    send_script = f"""
                    (async () => {{
                        try {{
                            // Check if number exists on WhatsApp
                            const isRegistered = await client.isRegisteredUser('{phone_number}@c.us');
                            if (!isRegistered) {{
                                console.log('NUMBER_NOT_ON_WHATSAPP');
                                return;
                            }}
                            
                            // Send the message
                            await client.sendMessage('{phone_number}@c.us', `{personalized_message}`);
                            console.log('SUCCESS');
                        }} catch(error) {{
                            console.log('ERROR: ' + error.message);
                        }}
                    }})();
                    """

                    # Write the script to a temporary file
                    temp_script_fd, temp_script_path = tempfile.mkstemp(
                        suffix='.js')
                    os.close(temp_script_fd)
                    with open(temp_script_path, 'w') as f:
                        f.write(send_script)

                    # Run the script with Node.js
                    process = subprocess.Popen(
                        ['node', temp_script_path],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True
                    )
                    stdout, stderr = process.communicate()

                    # Clean up the temporary script
                    os.remove(temp_script_path)

                    # Check the result
                    if 'SUCCESS' in stdout:
                        status = "Success"
                        sheet.cell(row=row, column=3).fill = success_fill
                    elif 'NUMBER_NOT_ON_WHATSAPP' in stdout:
                        status = "Number Doesn't Exist on WhatsApp"
                        sheet.cell(
                            row=row, column=3).fill = not_on_whatsapp_fill
                    else:
                        status = "Fail"
                        sheet.cell(row=row, column=3).fill = fail_fill

            except Exception as e:
                status = "Fail"
                sheet.cell(row=row, column=3).fill = fail_fill
                print(f"Error sending message to {phone_number}: {str(e)}")

            # Update the status in the Excel file
            sheet.cell(row=row, column=3).value = status

            # Save the workbook after each message to preserve progress
            workbook.save(file_path)

            # Add a random delay between messages (10-30 seconds)
            delay = random.randint(10, 30)
            time.sleep(delay)

        # Clean up temporary image file if it exists
        if temp_image_path and os.path.exists(temp_image_path):
            os.remove(temp_image_path)

    except Exception as e:
        print(f"Error processing bulk messages: {str(e)}")

# WebSocket event handlers


@socketio.on('connect')
def handle_connect():
    """Handle WebSocket connection."""
    global bot_connected, bot_process
    client_id = request.sid
    app.logger.info(f"Client connected: {client_id}")

    # Check if bot process is running
    bot_running = bot_process is not None and bot_process.poll() is None

    # Determine status
    status = 'connected' if bot_connected else 'disconnected'
    if bot_running and not bot_connected:
        status = 'connecting'

    # Send initial status
    emit('bot_status', {
        'connected': bot_connected,
        'status': status,
        'bot_running': bot_running
    })

    # Only send QR code info if the bot is in a connecting state
    # or if it's already connected
    if status == 'connecting' or bot_connected:
        qr_exists = os.path.exists('qr_code.png')
        emit('qr_code', {
            'exists': qr_exists,
            'qr_code_url': f'/get_qr_code?t={int(time.time())}' if qr_exists else None,
            'status': 'waiting_for_scan' if qr_exists else 'no_qr'
        })


@socketio.on('disconnect')
def handle_disconnect():
    app.logger.info(f"Client disconnected: {request.sid}")

# Background task for sending updates


def background_update_task():
    """Background task to periodically update clients with latest information."""
    global should_run_background_tasks, bot_connected, bot_process

    last_qr_status = False
    last_bot_status = bot_connected
    last_bot_running = bot_process is not None and bot_process.poll() is None

    while should_run_background_tasks:
        try:
            # Check if bot process is running
            bot_running = bot_process is not None and bot_process.poll() is None

            # Determine status
            status = 'connected' if bot_connected else 'disconnected'
            if bot_running and not bot_connected:
                status = 'connecting'

            # Check if QR code exists
            qr_exists = os.path.exists('qr_code.png')

            # Only emit QR code updates if the bot is in a connecting state
            # or if it's already connected
            if (status == 'connecting' or bot_connected) and qr_exists != last_qr_status:
                socketio.emit('qr_code', {
                    'exists': qr_exists,
                    'qr_code_url': f'/get_qr_code?t={int(time.time())}' if qr_exists else None,
                    'status': 'waiting_for_scan' if qr_exists else 'no_qr'
                })
                last_qr_status = qr_exists

            # Check bot status and emit if changed
            if bot_connected != last_bot_status or bot_running != last_bot_running:
                socketio.emit('bot_status', {
                    'connected': bot_connected,
                    'status': status,
                    'bot_running': bot_running
                })
                last_bot_status = bot_connected
                last_bot_running = bot_running

            time.sleep(1)
        except Exception as e:
            app.logger.error(f"Error in background task: {str(e)}")
            time.sleep(5)  # Wait longer if there's an error


def get_system_info():
    """Helper function to get system information"""
    # Get server time
    server_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Get system uptime
    try:
        uptime_seconds = time.time() - psutil.boot_time()
        uptime = str(timedelta(seconds=int(uptime_seconds)))
    except:
        uptime = "Unknown"

    # Get Node.js version if available
    try:
        node_version = subprocess.check_output(
            ['node', '--version']).decode().strip()
    except:
        node_version = "Not installed"

    # Get Python version
    python_version = f"{platform.python_version()} ({platform.python_implementation()})"

    return {
        "server_time": server_time,
        "uptime": uptime,
        "node_version": node_version,
        "python_version": python_version
    }


@app.route('/upload_image', methods=['POST'])
@login_required
def upload_image():
    if 'image' not in request.files:
        return jsonify({"message": "No image part", "success": False})

    file = request.files['image']
    if file.filename == '':
        return jsonify({"message": "No image selected", "success": False})

    keywords = request.form.get('keywords', '')
    if not keywords:
        return jsonify({"message": "Keywords are required", "success": False})

    # Split keywords by commas and strip whitespace
    keyword_list = [k.strip().lower()
                    for k in keywords.split(',') if k.strip()]

    if not keyword_list:
        return jsonify({"message": "At least one valid keyword is required", "success": False})

    if file and allowed_image_file(file.filename):
        # Generate a unique filename using uuid
        filename = secure_filename(file.filename)
        file_ext = filename.rsplit('.', 1)[1].lower()
        unique_filename = f"{str(uuid.uuid4())}.{file_ext}"

        # Create pics directory if it doesn't exist
        if not os.path.exists('pics'):
            os.makedirs('pics')

        # Save the file
        file_path = os.path.join('pics', unique_filename)
        file.save(file_path)

        # Update the image_keywords.json file
        try:
            # Load existing data
            if os.path.exists('image_keywords.json'):
                with open('image_keywords.json', 'r') as f:
                    try:
                        image_data = json.load(f)
                    except json.JSONDecodeError:
                        image_data = {}
            else:
                image_data = {}

            # Update data with new image
            for keyword in keyword_list:
                if keyword not in image_data:
                    image_data[keyword] = []

                # Add the image if not already present
                if unique_filename not in image_data[keyword]:
                    image_data[keyword].append(unique_filename)

            # Save the updated data
            with open('image_keywords.json', 'w') as f:
                json.dump(image_data, f, indent=2)

            return jsonify({
                "message": "Image uploaded successfully",
                "success": True,
                "filename": unique_filename,
                "keywords": keyword_list
            })

        except Exception as e:
            return jsonify({"message": f"Error updating image data: {str(e)}", "success": False})

    return jsonify({"message": "Invalid image file", "success": False})


@app.route('/get_images', methods=['GET'])
@login_required
def get_images():
    # Load image keywords data
    if os.path.exists('image_keywords.json'):
        with open('image_keywords.json', 'r') as f:
            try:
                image_data = json.load(f)
            except json.JSONDecodeError:
                image_data = {}
    else:
        image_data = {}

    # Get list of actual image files in the pics directory
    pic_files = []
    if os.path.exists('pics'):
        pic_files = [f for f in os.listdir('pics') if os.path.isfile(os.path.join('pics', f))
                     and allowed_image_file(f)]

    # Create a list of images with their keywords
    images = []
    for keyword, files in image_data.items():
        for file in files:
            if file in pic_files:  # Only include files that actually exist
                image_info = {
                    "filename": file,
                    "path": f"/pics/{file}",
                    "keywords": []
                }
                # Find all keywords associated with this file
                for k, f_list in image_data.items():
                    if file in f_list:
                        image_info["keywords"].append(k)

                # Only add each image once
                if not any(img["filename"] == file for img in images):
                    images.append(image_info)

    return jsonify({"images": images})


@app.route('/pics/<filename>')
@login_required
def serve_image(filename):
    return send_file(os.path.join('pics', filename))


@app.route('/delete_image', methods=['POST'])
@login_required
def delete_image():
    data = request.get_json()
    if not data or 'filename' not in data:
        return jsonify({"message": "No filename provided", "success": False})

    filename = data['filename']

    # Check if file exists
    file_path = os.path.join('pics', filename)
    if not os.path.exists(file_path):
        return jsonify({"message": "File not found", "success": False})

    try:
        # Remove the file
        os.remove(file_path)

        # Update the image_keywords.json file
        if os.path.exists('image_keywords.json'):
            with open('image_keywords.json', 'r') as f:
                try:
                    image_data = json.load(f)

                    # Remove the filename from all keyword lists
                    for keyword in image_data:
                        if filename in image_data[keyword]:
                            image_data[keyword].remove(filename)

                    # Remove any empty keyword entries
                    image_data = {k: v for k, v in image_data.items() if v}

                    # Save the updated data
                    with open('image_keywords.json', 'w') as f:
                        json.dump(image_data, f, indent=2)

                except json.JSONDecodeError:
                    pass

        return jsonify({
            "message": "Image deleted successfully",
            "success": True
        })

    except Exception as e:
        return jsonify({"message": f"Error deleting image: {str(e)}", "success": False})


def allowed_image_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS


@app.route('/get_ignore_list')
@login_required
def get_ignore_list():
    """Endpoint to get the numbers that are in the ignore list"""
    try:
        ignore_list_file = 'ignore_list.json'
        if os.path.exists(ignore_list_file):
            with open(ignore_list_file, 'r') as f:
                try:
                    data = f.read()
                    # Remove BOM and other potential invalid characters
                    if data and data[0] == '\ufeff':
                        data = data[1:]
                    # Clean the data of any non-printable characters
                    data = data.replace('\x00', '')
                    data = data.replace('/[^\x20-\x7E\r\n]/g', '')

                    # Handle empty or malformed file
                    if data.strip() == '':
                        return jsonify({"success": True, "ignored_numbers": []})

                    # Try to parse the JSON
                    ignore_list = json.loads(data)
                    return jsonify({"success": True, "ignored_numbers": ignore_list})
                except json.JSONDecodeError as e:
                    return jsonify({"success": False, "message": f"Error parsing ignore list: {str(e)}"})
        else:
            return jsonify({"success": True, "ignored_numbers": []})
    except Exception as e:
        return jsonify({"success": False, "message": f"Error getting ignore list: {str(e)}"})


# Function to broadcast ignore list changes to connected clients
def notify_ignore_list_change():
    """Broadcast to all clients that the ignore list has changed"""
    try:
        socketio.emit('ignore_list_updated')
    except Exception as e:
        app.logger.error(f"Error emitting ignore list update: {str(e)}")


@app.route('/notify_ignore_list_update', methods=['POST'])
def notify_ignore_list_update():
    """Endpoint to notify all connected clients that the ignore list has been updated"""
    try:
        notify_ignore_list_change()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)})


if __name__ == '__main__':
    app.logger.info("Starting the Flask application...")

    # Start background task in a separate thread
    background_thread = threading.Thread(target=background_update_task)
    background_thread.daemon = True
    background_thread.start()

    try:
        # Update the SocketIO run method to fix WebSocket issues
        # Note: We're specifically using gevent-websocket for the WebSocket transport
        # Create a middleware for the WebSocket handler
        from geventwebsocket.handler import WebSocketHandler
        from gevent.pywsgi import WSGIServer

        # Use gevent with WebSocketHandler
        server = WSGIServer(('0.0.0.0', 8080), app,
                            handler_class=WebSocketHandler)
        socketio.init_app(app)
        server.serve_forever()
    finally:
        # Signal background thread to stop
        should_run_background_tasks = False
        if background_thread.is_alive():
            background_thread.join(timeout=5)
